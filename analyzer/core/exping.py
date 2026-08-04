"""EXPING 로그(csv/xlsx) 형식 입출력 + pcap 의 ICMP echo 교환에서 로그 재구성.

EXPING 은 일본산 Windows ping 도구다. 6열(結果/日時/対象/ＩＰアドレス/ステータス/備考)
CSV 를 뽑고, 사용자는 그것을 엑셀로 열어 xlsx 로 보관한다. 이 모듈은 두 가지를 한다.

  1) 형식 입출력 — `read_csv()` / `write_csv()` / `write_xlsx()`
  2) pcap 재구성 — `extract_exchanges()` → `exchanges_to_rows()`

**왜 `analyzer.core.ping_matching` 을 쓰지 않나.** 그쪽은 대시보드용 손실 통계라
802.11 retry dedup, 단방향 흐름의 seq gap 추정 같은 추론이 들어간다. 여기서 필요한
것은 추론이 아니라 "echo request 1건 = 로그 1행, 캡처 순서 그대로"라는 EXPING 의
기록 규칙을 그대로 재현하는 것이다. 전제가 다르다. 또 이 모듈은 웹 파이프라인 없이
CLI 로만 쓰이므로 `extractor.extract_frames()` 의 무거운 WLAN 필드셋 대신 tshark 로
ICMP 최소 필드만 읽는다.

**재구성 규칙의 근거는 실측이다.** 시간동기화된 기준쌍(pcap ↔ exping csv, 10,433행)
에서 echo request 1건 ↔ 로그 1행이 정확히 성립했고, 무응답 23건의 인덱스가 ＮＧ 23행의
인덱스와 완전히 일치했다. 대상이 여러 개인 로그(3대상 회전)에서도 요청 시각 순 그대로가
로그 순서였다. 상세한 근거·한계는 `docs/EXPING.md` 참조.
"""

from __future__ import annotations

import csv
import datetime as dt
import math
import re
import subprocess
import sys
import tempfile
import threading
import zipfile
from dataclasses import dataclass
from pathlib import Path

# --------------------------------------------------------------------------
# 형식 상수
# --------------------------------------------------------------------------

#: EXPING CSV 머리글. 전각 문자다 — 'IP' 가 아니라 'ＩＰ'.
HEADER: tuple[str, ...] = ("結果", "日時", "対象", "ＩＰアドレス", "ステータス", "備考")

RESULT_OK = "ＯＫ"
RESULT_NG = "ＮＧ"
STATUS_TIMEOUT = "Request timed out"

#: 왕복시간 → 정수 ms 변환 보정값. `floor(왕복ms + 0.276)`.
#:
#: EXPING 이 쓰는 Windows IcmpSendEcho 의 측정구간이 와이어보다 조금 넓어서 생기는
#: 계통차다. 기준 표본 11,410건(2026-07-22 TEST1 응답 10,410 + 2026-07-21 TEST1
#: 잔존 1,000)에 합동 적합한 값이며 정수 ms 일치율 95.0%, 빗나간 515건 중 514건이
#: ±1 ms 다. 완전 일치는 원리상 불가능하다 — 같은 와이어 왕복시간(0.7 ms)이 어떤
#: 행에서는 0 ms, 어떤 행에서는 1 ms 로 기록된 사례가 실측에 있다.
#:
#: 실행마다 최적값이 0.28~0.44 ms 로 갈린다. 지연이 큰 캡처를 다룰 때는
#: `--rtt-offset` 으로 조정할 수 있다.
RTT_OFFSET_MS = 0.276

#: tshark 자식 프로세스 상한(초). 손상 캡처에서 멈추면 CLI 가 무기한 매달린다.
#: 실측 최대 133MB / 2시간 캡처가 분 단위라 넉넉하다. timesync-batch.py 와 같은 값.
CHILD_TIMEOUT = 3600

#: 응답으로 인정하는 상한(초). 이보다 늦게 온 echo reply 는 EXPING 도 타임아웃 처리한다.
#:
#: 실측: 2026-07-23 TEST14 에서 EXPING 이 ＮＧ 로 적은 31건의 실제 왕복시간이
#: 1,011~2,088 ms 였다. 상한을 1초로 두면 그 캡처의 손실 판정이 99.94% → 100% 가 된다.
DEFAULT_REPLY_TIMEOUT = 1.0

#: 엑셀 시트명 상한. 넘으면 엑셀이 자른다.
SHEET_NAME_LIMIT = 31

#: 기준 파일의 열 너비 (A~F).
COLUMN_WIDTHS: tuple[float, ...] = (7.5, 15.25, 11.75, 15.25, 18.0, 7.5)

TABLE_STYLE = "TableStyleMedium7"
DEFAULT_FONT_NAME = "맑은 고딕"
DEFAULT_ROW_HEIGHT = 16.5
#: 시트1(표 시트)은 시각을, 시트2는 기준 파일이 그렇듯 엑셀 기본 날짜서식을 쓴다.
TABLE_SHEET_TIME_FORMAT = "h:mm:ss;@"
PLAIN_SHEET_TIME_FORMAT = "m/d/yy h:mm"

_CSV_TIME_FORMAT = "%Y-%m-%d %H:%M:%S"

#: 시트에 쓰는 한 행 — HEADER 와 같은 순서. 빈 칸은 None.
SheetRow = tuple[
    "str | None", dt.datetime, "str | None", "str | None", "str | None", "str | None"
]


@dataclass(frozen=True)
class Exchange:
    """echo request 한 건과 그 응답.

    `rtt` 가 None 이면 인정 시간 안에 응답이 없었다는 뜻이다.
    """

    time: float
    target: str
    rtt: float | None

    @property
    def answered(self) -> bool:
        return self.rtt is not None


# --------------------------------------------------------------------------
# 행 만들기
# --------------------------------------------------------------------------


def format_status(rtt: float | None, offset_ms: float = RTT_OFFSET_MS) -> str:
    """ステータス 열 문자열. 응답 없으면 타임아웃 문구."""
    if rtt is None:
        return STATUS_TIMEOUT
    return "Time:%6dms" % max(0, math.floor(rtt * 1000 + offset_ms))


def local_stamp(epoch: float, tz: dt.tzinfo | None = None) -> dt.datetime:
    """epoch 을 표시용 naive datetime 으로. `tz` 생략 시 시스템 지역시각.

    엑셀은 타임존 붙은 datetime 을 거부하므로 변환 후 tzinfo 를 떼어낸다.
    `os.environ["TZ"]` + `time.tzset()` 을 쓰지 않는 이유: tzset 은 POSIX 전용이라
    Windows 에서 죽고, 프로세스 전역 상태를 건드려 다른 코드에 샌다.
    """
    if tz is None:
        return dt.datetime.fromtimestamp(epoch)
    return dt.datetime.fromtimestamp(epoch, tz).replace(tzinfo=None)


def exchanges_to_rows(
    exchanges: list[Exchange],
    offset_ms: float = RTT_OFFSET_MS,
    tz: dt.tzinfo | None = None,
) -> list[SheetRow]:
    """교환 목록을 시트 행으로. 日時 는 요청 시각을 초 단위로 절삭한다."""
    rows: list[SheetRow] = []
    for ex in exchanges:
        stamp = local_stamp(ex.time, tz).replace(microsecond=0)
        rows.append(
            (
                RESULT_OK if ex.answered else RESULT_NG,
                stamp,
                ex.target,
                ex.target if ex.answered else None,
                format_status(ex.rtt, offset_ms),
                None,
            )
        )
    return rows


# --------------------------------------------------------------------------
# CSV 입출력
# --------------------------------------------------------------------------


def read_csv(path: str | Path) -> list[SheetRow]:
    """EXPING CSV 를 시트 행으로 읽는다. ステータス 는 손대지 않고 그대로 옮긴다.

    'Time: N ms' / 'Request timed out' 말고도 'Destination host unreachable',
    'Unknown Error' 같은 값이 실제로 나온다. 해석하지 않는 이유다.
    """
    # newline="" 은 csv 모듈 규약 — 필드 안에 개행이 들어와도 깨지지 않는다.
    with open(path, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        raise ValueError(f"빈 파일: {path}")
    if tuple(rows[0]) != HEADER:
        raise ValueError(f"EXPING 머리글이 아니다: {path}\n  읽은 값: {rows[0]}")
    out: list[SheetRow] = []
    for lineno, r in enumerate(rows[1:], start=2):
        if len(r) != len(HEADER):
            raise ValueError(f"{path}:{lineno} 열 개수가 {len(HEADER)}이 아니다: {r}")
        try:
            stamp = dt.datetime.strptime(r[1], _CSV_TIME_FORMAT)
        except ValueError as exc:
            raise ValueError(f"{path}:{lineno} 日時 형식 오류: {r[1]!r}") from exc
        out.append((r[0] or None, stamp, r[2] or None, r[3] or None, r[4] or None, r[5] or None))
    return out


def write_csv(path: str | Path, rows: list[SheetRow]) -> None:
    """EXPING 이 뽑는 것과 같은 바이트 형식: UTF-8 BOM, CRLF, 전 필드 인용."""
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writerow(HEADER)
        for r in rows:
            w.writerow([r[0] or "", r[1].strftime(_CSV_TIME_FORMAT), *(v or "" for v in r[2:])])


# --------------------------------------------------------------------------
# XLSX 출력
# --------------------------------------------------------------------------


def sheet_names(base: str) -> tuple[str, str]:
    """기준 파일과 같은 두 시트 이름 (표 시트, 원본 시트).

    엑셀은 시트 복사본에 " (2)" 를 붙이고 31자를 넘으면 **원래 이름 쪽**을 자른다.
    기준 파일이 'FXA3000_RAIL_TEST1(1514_1534)' → 'FXA3000_RAIL_TEST1(1514_153 (2)'
    인 것이 그 증거다.
    """
    suffix = " (2)"
    if len(base) + len(suffix) <= SHEET_NAME_LIMIT:
        dup = base + suffix
    else:
        dup = base[: SHEET_NAME_LIMIT - len(suffix)] + suffix
    return dup, base[:SHEET_NAME_LIMIT]


def table_name(base: str) -> str:
    """엑셀 표 이름 — 영숫자/밑줄만 허용되고 끝의 밑줄은 엑셀이 떼어낸다.

    단어 문자가 하나도 없는 이름(예: `---`)이면 빈 문자열이 되어 엑셀이 거부하므로
    대체 이름을 쓴다.
    """
    return re.sub(r"\W", "_", base).strip("_") or "ExpingTable"


def write_xlsx(
    path: str | Path,
    rows: list[SheetRow],
    base: str,
    theme_from: str | Path | None = None,
) -> None:
    """기준 파일과 같은 서식의 2시트 워크북으로 쓴다.

    `theme_from` 에 기준 xlsx 를 주면 테마(표 스타일 색)를 그대로 가져온다. 생략하면
    openpyxl 기본 테마라 표 색이 달라 보인다.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Color, Font
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - 선택 의존성
        raise RuntimeError(
            "xlsx 출력에는 openpyxl 이 필요하다: pip install -r requirements-exping.txt"
        ) from exc

    wb = openpyxl.Workbook()
    # 기준 파일과 같은 기본 글꼴. openpyxl 기본은 Calibri 라 그냥 두면 눈에 띈다.
    # `_fonts` 는 비공개 속성이다 — 기본 글꼴을 바꾸는 공개 API 가 없다. openpyxl 이
    # 내부 구조를 바꾸면 글꼴만 포기하고 계속 간다 (데이터에는 영향 없음).
    # requirements-exping.txt 에서 메이저 버전을 묶어 급변을 막는다.
    try:
        wb._fonts[0] = Font(
            name=DEFAULT_FONT_NAME,
            sz=11,
            color=Color(theme=1, type="theme"),
            family=2,
            charset=129,
            scheme="minor",
        )
    except (AttributeError, IndexError, TypeError) as exc:  # pragma: no cover - 방어
        print(f"[!] 기본 글꼴 설정 실패, 그대로 진행한다 ({exc})", file=sys.stderr)

    if theme_from:
        # 테마는 표 색만 좌우한다. 못 읽어도 중단하지 않고 기본 테마로 간다.
        try:
            with zipfile.ZipFile(theme_from) as z:
                wb.loaded_theme = z.read("xl/theme/theme1.xml")
        except (OSError, zipfile.BadZipFile, KeyError) as exc:
            print(f"[!] 테마를 못 읽어 기본 테마로 진행한다: {theme_from} ({exc})", file=sys.stderr)

    center = Alignment(vertical="center")
    dup_name, plain_name = sheet_names(base)

    def fill(ws, time_format: str) -> None:
        ws.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT
        ws.append(list(HEADER))
        for r in rows:
            ws.append(list(r))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(HEADER)):
            for c in row:
                c.alignment = center
        for c in ws["B"][1:]:
            c.number_format = time_format

    ws = wb.active
    ws.title = dup_name
    fill(ws, TABLE_SHEET_TIME_FORMAT)
    # fill() 의 루프는 ws["B"][1:] 이라 B1 을 건너뛴다 — 머리글 셀 서식은 여기서만 정한다.
    # (기준 파일과 맞추는 부분이니 "중복"으로 보고 지우면 조용히 어긋난다.)
    ws["B1"].number_format = TABLE_SHEET_TIME_FORMAT
    for col, width in zip("ABCDEF", COLUMN_WIDTHS):
        ws.column_dimensions[col].width = width
    ws.column_dimensions["B"].number_format = TABLE_SHEET_TIME_FORMAT
    table = Table(displayName=table_name(base), ref=f"A1:F{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name=TABLE_STYLE, showRowStripes=True)
    ws.add_table(table)

    fill(wb.create_sheet(plain_name), PLAIN_SHEET_TIME_FORMAT)
    wb.save(path)


# --------------------------------------------------------------------------
# pcap 에서 교환 뽑기
# --------------------------------------------------------------------------

#: tshark 로 읽는 최소 필드. 순서가 곧 파싱 순서다.
TSHARK_FIELDS: tuple[str, ...] = (
    "frame.time_epoch",
    "ip.src",
    "ip.dst",
    "icmp.type",
    "icmp.ident",
    "icmp.seq",
    # 802.11 헤더가 붙은 프레임에만 값이 있다. 유선 캡처에서는 늘 빈 칸이다.
    # 파일 단위 encap 은 인터페이스가 여럿인 pcapng 에서 틀린 답을 준다 — 프레임 단위로 본다.
    "wlan.fc.type",
)

ICMP_ECHO_REQUEST = "8"
ICMP_ECHO_REPLY = "0"

#: 파싱된 ICMP 프레임 한 줄 — TSHARK_FIELDS 와 같은 순서.
IcmpFrame = tuple[float, str, str, str, str, str, str]


def build_tshark_cmd(pcap: str | Path, tshark: str = "tshark") -> list[str]:
    cmd = [tshark, "-r", str(pcap), "-Y", "icmp.type==8 || icmp.type==0", "-T", "fields"]
    for f in TSHARK_FIELDS:
        cmd += ["-e", f]
    return cmd


def parse_icmp_line(line: str) -> IcmpFrame | None:
    """tshark 탭 출력 한 줄. 필드가 빠졌거나 시각을 못 읽으면 None."""
    f = line.rstrip("\n").split("\t")
    if len(f) != len(TSHARK_FIELDS):
        return None
    try:
        epoch = float(f[0])
    except ValueError:
        return None
    return (epoch, f[1], f[2], f[3], f[4], f[5], f[6])


def parse_icmp_tsv(text: str) -> list[IcmpFrame]:
    """tshark 탭 출력 파싱. 필드가 빠졌거나 시각을 못 읽는 줄은 버린다."""
    return [f for f in (parse_icmp_line(ln) for ln in text.splitlines()) if f is not None]


def pick_sender(frames: list[IcmpFrame]) -> str:
    """echo request 를 가장 많이 보낸 호스트를 EXPING 실행 PC 로 본다."""
    counts: dict[str, int] = {}
    for _, src, _dst, typ, _ident, _seq, _wlan in frames:
        if typ == ICMP_ECHO_REQUEST:
            counts[src] = counts.get(src, 0) + 1
    if not counts:
        raise ValueError("ICMP echo request 가 없다")
    return max(counts.items(), key=lambda kv: kv[1])[0]


def pair_exchanges(
    frames: list[IcmpFrame], sender: str, timeout: float = DEFAULT_REPLY_TIMEOUT
) -> list[Exchange]:
    """요청 시각 순으로 (요청, 응답) 을 짝짓는다.

    응답은 (ident, seq, 상대IP) 로 찾는다. 대상이 여러 개면 EXPING 은 대상을 번갈아
    보내고 로그도 그 순서 그대로다 — 그래서 대상별로 나누지 않고 시각 순 한 줄로 낸다.

    **짝지은 응답을 목록에서 빼지 않는다.** 같은 키가 창 안에 두 번 나오면 같은 응답이
    여러 요청에 매핑될 수 있는데, 그래도 그냥 두는 이유:

    - seq 순환은 못 온다. icmp.seq 는 16비트라 관측된 8.5 ping/s 로 2시간 넘게 걸리는데,
      응답 인정 창은 1초다.
    - 창 안에서 키가 겹치는 경우는 모니터 캡처의 802.11 retry 사본뿐이다. 그때는 문제가
      "응답 재사용"이 아니라 **요청 행 자체가 중복**인 것이라(무선 11,224 대 유선 9,296),
      응답을 소비하도록 바꾸면 OK 9,467 → 7,967 로 되레 틀려진다.
    - 지원 대상인 유선 캡처에서는 중복 키가 0건이라 아무 차이가 없다.
    """
    replies: dict[tuple[str, str, str], list[float]] = {}
    for epoch, src, dst, typ, ident, seq, _wlan in frames:
        if typ == ICMP_ECHO_REPLY and dst == sender:
            replies.setdefault((ident, seq, src), []).append(epoch)
    for v in replies.values():
        v.sort()

    requests = sorted(
        (f for f in frames if f[3] == ICMP_ECHO_REQUEST and f[1] == sender),
        key=lambda f: f[0],
    )
    out: list[Exchange] = []
    for epoch, _src, dst, _typ, ident, seq, _wlan in requests:
        cand = [x for x in replies.get((ident, seq, dst), ()) if epoch <= x <= epoch + timeout]
        out.append(Exchange(epoch, dst, (cand[0] - epoch) if cand else None))
    return out


def drop_trailing_unanswered(exchanges: list[Exchange]) -> tuple[list[Exchange], int]:
    """끝에 붙은 무응답 요청을 떼어낸다. 손실인지 알 수 없는 구간이라 ＮＧ 오탐이 된다.

    실측: 2026-07-21 TEST7 캡처는 마지막 요청 0.000초 뒤에 끝났고, 그 행은 원본 EXPING
    로그에 ＯＫ 1 ms 로 남아 있다. 응답이 없었던 게 아니라 캡처가 못 잡은 것이다.
    """
    kept = list(exchanges)
    dropped = 0
    while kept and not kept[-1].answered:
        kept.pop()
        dropped += 1
    return kept, dropped


def count_wireless_requests(frames: list[IcmpFrame], sender: str) -> tuple[int, int]:
    """(sender 가 보낸 echo request 수, 그중 802.11 프레임에서 온 수).

    무선(모니터) 캡처를 넣으면 도구는 그냥 돌아가지만 결과가 조용히 틀린다 —
    모니터가 못 들은 프레임이 전부 손실로 계산된다. 실측(2026-07-21 TEST1):
    유선 9,296행 / ＮＧ 15 / 0.16% 대 무선 11,224행 / ＮＧ 1,757 / **15.65%**.

    파일 단위 encap 이 아니라 프레임 단위로 세는 이유: 인터페이스가 여럿인 pcapng
    (유선+무선을 함께 뜬 것, mergecap 결과)에서는 파일 단위 판정이 무의미하고
    같은 ping 이 두 번 세어진다.
    """
    total = wireless = 0
    for _epoch, src, _dst, typ, _ident, _seq, wlan in frames:
        if typ == ICMP_ECHO_REQUEST and src == sender:
            total += 1
            wireless += bool(wlan)
    return total, wireless


def extract_icmp_frames(
    pcap: str | Path,
    tshark: str = "tshark",
    child_timeout: float | None = CHILD_TIMEOUT,
    cancel_event: "threading.Event | None" = None,
    warnings_out: "list[str] | None" = None,
) -> list[IcmpFrame]:
    """pcap 에서 ICMP echo request/reply 프레임만 스트리밍 추출한다.

    `extract_exchanges` 의 하위 레이어다 — tshark 서브프로세스 실행·부분 실패
    허용·자식 타임아웃·stderr 임시파일 캡처만 담당하고, sender 선정/무선 가드/
    짝짓기는 하지 않는다. `analyzer.core.wired_ping.build_ground_truth` 가 시간/IP
    필터가 적용된 부분집합에서 sender 를 다시 고르기 위해(전체 pcap 최다 요청자가
    아니라 필터 구간의 최다 요청자를 sender 로 써야 한다) 이 레이어를 직접
    재사용한다 — `extract_exchanges` 를 거치면 sender 가 전체 pcap 기준으로
    이미 확정돼 버린다.

    tshark 출력을 한 줄씩 걸러 담는다 — 프레임이 수백만인 캡처에서 출력 전체를
    메모리에 올리지 않기 위해서다. `check=True` 를 쓰지 않는 이유는
    `timesync.extract_ntp_responses()` 와 같다: 종료코드를 직접 보고 사람이 읽을 수
    있는 오류를 낸다. 프레임을 건졌다면 비정상 종료라도 그 결과는 쓴다.

    stderr 는 파이프가 아니라 임시 파일로 받는다. 파이프로 받으면 경고가 많은 캡처에서
    stderr 버퍼(보통 64KB)가 차는 순간 tshark 가 쓰기에서 멈추고, 부모는 stdout 만
    읽고 있어 EOF 가 오지 않아 서로 영원히 기다린다.

    `child_timeout` 은 tshark 상한(초)이다. `None`(또는 0)이면 상한을 걸지 않는다 —
    무기한 대기를 감수한다는 뜻이니 의도한 경우에만 쓴다.

    `cancel_event` 를 주면 그 이벤트가 set 되는 즉시 tshark 자식을 종료하고
    `InterruptedError` 를 낸다(무선 추출 `extractor.extract_frames` 의 취소 계약과
    같다 — 웹에서 /api/cancel 이 성공을 보고했는데 자식은 child_timeout 까지 살아
    임시파일과 pcap 핸들을 쥐고 있는 일을 막는다). 부분 프레임으로 계속하지 않는
    이유: 취소 시점까지 읽은 것만으로 집계하면 손실률이 조용히 틀린다. 예외를
    새로 쓰는 이유: 기존 `ValueError`/`TimeoutError` 계약을 건드리지 않기 위함.
    미전달(기본 `None`)이면 감시 스레드를 아예 만들지 않아 동작이 불변이다.

    `warnings_out` 리스트를 주면 부분 실패(비정상 종료했지만 프레임은 건진 경우)
    경고를 stderr 와 **같은 내용**으로 거기에도 담는다. stderr 만으로는 웹 경로가
    그 사실을 알 길이 없어, 잘린/손상 pcap 이 아무 경고 없이 "성공한 ground
    truth" 로 게시된다(손실률이 실제보다 낮게 나온다). stderr 출력은 그대로
    유지되므로 CLI 동작은 불변이다.
    """
    if cancel_event is not None and cancel_event.is_set():
        # 이미 취소된 뒤라면 자식을 띄우지도 않는다.
        raise InterruptedError("취소됨")

    frames: list[IcmpFrame] = []
    with tempfile.TemporaryFile("w+", encoding="utf-8", errors="replace") as errf:
        proc = subprocess.Popen(
            build_tshark_cmd(pcap, tshark),
            stdout=subprocess.PIPE,
            stderr=errf,
            text=True,
        )
        # 파이프 읽기에서 블록되면 루프 안의 시각 검사는 영영 실행되지 않는다.
        # 밖에서 죽여야 한다. 외부 kill 과 구분하려고 발동 여부를 따로 남긴다.
        timed_out: list[bool] = []
        cancelled: list[bool] = []

        def _kill_child() -> None:
            timed_out.append(True)
            proc.kill()

        def _cancel_watcher() -> None:
            # extractor._cancel_watcher 와 같은 패턴: 부모가 파이프 읽기에서 막혀
            # 있으므로 감시 스레드가 자식을 죽이고 stdout 을 닫아 읽기를 깨운다.
            while proc.poll() is None:
                if cancel_event.wait(0.05):
                    cancelled.append(True)
                    try:
                        proc.terminate()
                        proc.wait(timeout=1)
                    except subprocess.TimeoutExpired:
                        proc.kill()
                    except OSError:
                        pass
                    try:
                        if proc.stdout is not None:
                            proc.stdout.close()
                    except OSError:
                        pass
                    return

        killer = threading.Timer(child_timeout, _kill_child) if child_timeout else None
        if killer is not None:
            killer.daemon = True
            killer.start()
        watcher = (
            threading.Thread(target=_cancel_watcher, daemon=True)
            if cancel_event is not None
            else None
        )
        if watcher is not None:
            watcher.start()
        try:
            for line in proc.stdout or ():
                parsed = parse_icmp_line(line)
                if parsed is not None:
                    frames.append(parsed)
            returncode = proc.wait()
        except (ValueError, OSError):
            # 취소 감시 스레드가 stdout 을 닫으면 읽기가 여기로 떨어진다.
            # 취소가 아니면 원래대로 올려보낸다.
            if not cancelled:
                proc.kill()
                proc.wait()
                raise
            returncode = proc.wait()
        except BaseException:
            # 여기서 안 죽이면 tshark 가 고아로 남아 pcap 파일 핸들을 계속 쥔다.
            proc.kill()
            proc.wait()
            raise
        finally:
            if killer is not None:
                killer.cancel()
            if watcher is not None:
                watcher.join(timeout=1)
            if proc.stdout is not None:
                proc.stdout.close()
        errf.seek(0)
        err = errf.read()

    if cancelled or (cancel_event is not None and cancel_event.is_set()):
        raise InterruptedError("취소됨")

    if timed_out:
        raise TimeoutError(f"tshark 가 {child_timeout:g}초 안에 끝나지 않아 중단했다: {pcap}")

    if returncode != 0:
        detail = err.strip().splitlines()
        last = detail[-1] if detail else "출력 없음"
        if not frames:
            raise ValueError(f"tshark 가 실패했다 (exit {returncode}): {last}")
        # 프레임을 건졌어도 비정상 종료면 캡처가 중간에 끊겼을 수 있다. 조용히 넘기면
        # 못 읽은 요청만큼 손실률이 실제보다 **낮게** 나오는데 사용자는 알 길이 없다.
        # 문구를 한 번만 만들어 stderr(CLI)와 warnings_out(웹) 양쪽에 같은 내용을
        # 흘린다 — 두 경로가 갈라져 한쪽만 갱신되는 일을 막는다.
        partial_msgs = [
            f"tshark 가 exit {returncode} 로 끝났다 — 결과가 일부일 수 있다: {last}",
            f"읽은 ICMP 프레임 {len(frames):,}개로 계속한다. "
            "손실률이 실제보다 낮게 나올 수 있으니 캡처 무결성을 확인하라.",
        ]
        for msg in partial_msgs:
            print(f"[!] {msg}", file=sys.stderr)
        if warnings_out is not None:
            warnings_out.extend(partial_msgs)
    return frames


def extract_exchanges(
    pcap: str | Path,
    tshark: str = "tshark",
    sender: str | None = None,
    timeout: float = DEFAULT_REPLY_TIMEOUT,
    child_timeout: float | None = CHILD_TIMEOUT,
    allow_wireless: bool = False,
) -> tuple[list[Exchange], str]:
    """pcap 을 읽어 (교환 목록, 송신 호스트) 를 낸다.

    프레임 추출(서브프로세스 실행·부분 실패 허용·자식 타임아웃)은
    `extract_icmp_frames` 에 위임한다 — 이 함수는 그 위에 sender 선정(전체 pcap
    최다 요청자)·무선 캡처 가드·요청↔응답 짝짓기만 얹는다. 공개 시그니처·예외
    종류·stderr 메시지·CLI 동작은 리팩토링 전과 동일하다(tests/test_exping.py 전체
    그린으로 검증).
    """
    frames = extract_icmp_frames(pcap, tshark, child_timeout)
    src = sender or pick_sender(frames)

    total, wireless = count_wireless_requests(frames, src)
    if wireless and not allow_wireless:
        if wireless == total:
            raise ValueError(
                f"무선(802.11) 캡처다 — 유선 캡처를 넣어라: {pcap}\n"
                "  모니터가 못 들은 프레임이 전부 손실로 계산돼 손실률이 크게 부풀려진다 "
                "(실측 0.16% 대 15.65%).\n"
                "  그래도 진행하려면 --allow-wireless 를 붙여라."
            )
        print(
            f"[!] echo request {total:,}건 중 {wireless:,}건이 802.11 프레임이다 — "
            "인터페이스가 여럿인 캡처로 보인다.",
            file=sys.stderr,
        )
        print(
            "[!] 같은 ping 이 두 번 세어져 행 수와 손실률이 왜곡될 수 있다. "
            "유선 인터페이스만 담긴 캡처를 쓰는 편이 낫다.",
            file=sys.stderr,
        )

    return pair_exchanges(frames, src, timeout), src
